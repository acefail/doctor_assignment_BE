from openpyxl import load_workbook
def readData(sheetName):
    data_test = load_workbook("./dataSheet.xlsx")
    values = []
    s = data_test[sheetName]
    for row in s.iter_rows(min_row=2, values_only=True):  # Assuming your data starts from the second row
        values.append(row)      
    print(values)
    return values
LINK_URL = "http://localhost:3000/"